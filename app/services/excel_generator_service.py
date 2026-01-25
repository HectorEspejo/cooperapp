import os
from datetime import datetime
from decimal import Decimal
from io import BytesIO
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.project import Project, Financiador
from app.models.expense import Expense, EstadoGasto, UbicacionGasto
from app.models.transfer import Transfer
from app.models.budget import ProjectBudgetLine
from app.models.report import TipoInforme


# Funder colors for Excel headers
FUNDER_COLORS = {
    Financiador.aacid: "006633",          # AACID - Green
    Financiador.aecid: "C41E3A",          # AECID - Red
    Financiador.diputacion_malaga: "003366",  # DIPU - Navy blue
    Financiador.ayuntamiento_malaga: "8B1E3F",  # AYTO - Brand color
}


class ExcelGeneratorService:
    def __init__(self, db: Session):
        self.db = db

    def generate_report(
        self,
        project: Project,
        tipo: TipoInforme,
        periodo: str | None = None,
    ) -> tuple[BytesIO, str]:
        """Generate an Excel report and return the buffer and filename."""
        generators = {
            TipoInforme.cuenta_justificativa: self._generate_cuenta_justificativa,
            TipoInforme.ejecucion_presupuestaria: self._generate_ejecucion_presupuestaria,
            TipoInforme.relacion_transferencias: self._generate_relacion_transferencias,
            TipoInforme.ficha_proyecto: self._generate_ficha_proyecto,
            TipoInforme.informe_tecnico_mensual: self._generate_informe_tecnico,
            TipoInforme.informe_economico: self._generate_informe_economico,
        }

        generator = generators.get(tipo)
        if not generator:
            raise ValueError(f"Tipo de informe no soportado: {tipo}")

        return generator(project, periodo)

    def generate_pack(
        self,
        project: Project,
        tipos: list[TipoInforme] | None = None,
    ) -> tuple[BytesIO, str]:
        """Generate a ZIP pack with multiple reports."""
        if tipos is None:
            tipos = [
                TipoInforme.cuenta_justificativa,
                TipoInforme.ejecucion_presupuestaria,
                TipoInforme.relacion_transferencias,
            ]

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        zip_filename = f"pack_justificacion_{project.codigo_contable}_{timestamp}.zip"

        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for tipo in tipos:
                try:
                    excel_buffer, excel_filename = self.generate_report(project, tipo)
                    zip_file.writestr(excel_filename, excel_buffer.getvalue())
                except Exception:
                    # Skip reports that fail to generate
                    pass

        zip_buffer.seek(0)
        return zip_buffer, zip_filename

    # === Common Styles ===

    def _get_header_fill(self, project: Project) -> PatternFill:
        """Get the header fill color based on funder."""
        color = FUNDER_COLORS.get(project.financiador, "8B1E3F")
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    def _get_header_font(self) -> Font:
        """Get the header font style."""
        return Font(bold=True, color="FFFFFF", size=11)

    def _get_title_font(self) -> Font:
        """Get the title font style."""
        return Font(bold=True, size=14)

    def _get_thin_border(self) -> Border:
        """Get thin border for data cells."""
        thin = Side(style="thin", color="000000")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def _add_project_header(self, ws, project: Project, title: str):
        """Add common project header to worksheet."""
        # Title
        ws["A1"] = title
        ws["A1"].font = self._get_title_font()
        ws.merge_cells("A1:G1")

        # Project info
        ws["A3"] = "Codigo:"
        ws["B3"] = project.codigo_contable
        ws["A4"] = "Titulo:"
        ws["B4"] = project.titulo
        ws.merge_cells("B4:G4")
        ws["A5"] = "Financiador:"
        ws["B5"] = project.financiador.value
        ws["D5"] = "Generado:"
        ws["E5"] = datetime.now().strftime("%d/%m/%Y %H:%M")

        # Bold labels
        for cell in ["A3", "A4", "A5", "D5"]:
            ws[cell].font = Font(bold=True)

        return 7  # Return next available row

    def _auto_column_widths(self, ws, min_width: int = 10, max_width: int = 50):
        """Auto-adjust column widths based on content."""
        for column_cells in ws.columns:
            length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in column_cells
            )
            adjusted_width = min(max(length + 2, min_width), max_width)
            column_letter = get_column_letter(column_cells[0].column)
            ws.column_dimensions[column_letter].width = adjusted_width

    # === Report Generators ===

    def _generate_cuenta_justificativa(
        self, project: Project, periodo: str | None = None
    ) -> tuple[BytesIO, str]:
        """Generate Cuenta Justificativa report - Expense list."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Cuenta Justificativa"

        start_row = self._add_project_header(ws, project, "CUENTA JUSTIFICATIVA")

        # Column headers
        headers = [
            "Fecha",
            "Concepto",
            "Expedidor",
            "Partida",
            "Financiador",
            "Importe EUR",
            "Ubicacion",
            "Estado",
        ]
        header_fill = self._get_header_fill(project)
        header_font = self._get_header_font()
        border = self._get_thin_border()

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # Data rows - only validated/justified expenses
        row = start_row + 1
        total_importe = Decimal("0")

        expenses = sorted(project.expenses, key=lambda e: e.fecha_factura)
        for expense in expenses:
            if expense.estado not in (EstadoGasto.validado, EstadoGasto.justificado):
                continue

            ws.cell(row=row, column=1, value=expense.fecha_factura.strftime("%d/%m/%Y")).border = border
            ws.cell(row=row, column=2, value=expense.concepto).border = border
            ws.cell(row=row, column=3, value=expense.expedidor).border = border
            ws.cell(row=row, column=4, value=expense.budget_line.name if expense.budget_line else "").border = border
            ws.cell(row=row, column=5, value=expense.financiado_por).border = border

            importe_cell = ws.cell(row=row, column=6, value=float(expense.cantidad_imputable))
            importe_cell.number_format = '#,##0.00 "€"'
            importe_cell.border = border

            ubicacion = "España" if expense.ubicacion == UbicacionGasto.espana else "Terreno"
            ws.cell(row=row, column=7, value=ubicacion).border = border
            ws.cell(row=row, column=8, value=expense.estado.value.replace("_", " ").title()).border = border

            total_importe += expense.cantidad_imputable
            row += 1

        # Total row
        row += 1
        ws.cell(row=row, column=5, value="TOTAL:").font = Font(bold=True)
        total_cell = ws.cell(row=row, column=6, value=float(total_importe))
        total_cell.font = Font(bold=True)
        total_cell.number_format = '#,##0.00 "€"'

        self._auto_column_widths(ws)

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"cuenta_justificativa_{project.codigo_contable}_{timestamp}.xlsx"

        return buffer, filename

    def _generate_ejecucion_presupuestaria(
        self, project: Project, periodo: str | None = None
    ) -> tuple[BytesIO, str]:
        """Generate Ejecucion Presupuestaria report - Budget execution table."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Ejecucion Presupuestaria"

        start_row = self._add_project_header(ws, project, "EJECUCION PRESUPUESTARIA")

        # Column headers
        headers = [
            "Partida",
            "Aprobado",
            "Ejec. España",
            "Ejec. Terreno",
            "Total Ejecutado",
            "Diferencia",
            "% Ejecucion",
        ]
        header_fill = self._get_header_fill(project)
        header_font = self._get_header_font()
        border = self._get_thin_border()

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # Data rows
        row = start_row + 1
        totals = {
            "aprobado": Decimal("0"),
            "espana": Decimal("0"),
            "terreno": Decimal("0"),
            "ejecutado": Decimal("0"),
        }

        for budget_line in project.budget_lines:
            # Calculate execution by location
            espana = sum(
                e.cantidad_imputable
                for e in budget_line.expenses
                if e.ubicacion == UbicacionGasto.espana
                and e.estado in (EstadoGasto.validado, EstadoGasto.justificado)
            )
            terreno = sum(
                e.cantidad_imputable
                for e in budget_line.expenses
                if e.ubicacion == UbicacionGasto.terreno
                and e.estado in (EstadoGasto.validado, EstadoGasto.justificado)
            )
            ejecutado = espana + terreno
            aprobado = budget_line.aprobado or Decimal("0")
            diferencia = aprobado - ejecutado
            porcentaje = (ejecutado / aprobado * 100) if aprobado > 0 else Decimal("0")

            ws.cell(row=row, column=1, value=budget_line.name).border = border

            for col, value in [(2, aprobado), (3, espana), (4, terreno), (5, ejecutado), (6, diferencia)]:
                cell = ws.cell(row=row, column=col, value=float(value))
                cell.number_format = '#,##0.00 "€"'
                cell.border = border

            pct_cell = ws.cell(row=row, column=7, value=float(porcentaje) / 100)
            pct_cell.number_format = "0.00%"
            pct_cell.border = border

            totals["aprobado"] += aprobado
            totals["espana"] += espana
            totals["terreno"] += terreno
            totals["ejecutado"] += ejecutado
            row += 1

        # Total row
        row += 1
        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)

        diferencia_total = totals["aprobado"] - totals["ejecutado"]
        pct_total = (
            totals["ejecutado"] / totals["aprobado"] * 100
            if totals["aprobado"] > 0
            else Decimal("0")
        )

        for col, value in [
            (2, totals["aprobado"]),
            (3, totals["espana"]),
            (4, totals["terreno"]),
            (5, totals["ejecutado"]),
            (6, diferencia_total),
        ]:
            cell = ws.cell(row=row, column=col, value=float(value))
            cell.font = Font(bold=True)
            cell.number_format = '#,##0.00 "€"'

        pct_cell = ws.cell(row=row, column=7, value=float(pct_total) / 100)
        pct_cell.font = Font(bold=True)
        pct_cell.number_format = "0.00%"

        self._auto_column_widths(ws)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"ejecucion_presupuestaria_{project.codigo_contable}_{timestamp}.xlsx"

        return buffer, filename

    def _generate_relacion_transferencias(
        self, project: Project, periodo: str | None = None
    ) -> tuple[BytesIO, str]:
        """Generate Relacion de Transferencias report."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Transferencias"

        start_row = self._add_project_header(ws, project, "RELACION DE TRANSFERENCIAS")

        # Column headers
        headers = [
            "Nº",
            "F. Peticion",
            "F. Emision",
            "F. Recepcion",
            "Importe EUR",
            "Gastos",
            "Neto EUR",
            "Tipo Cambio",
            "Moneda Local",
            "Importe Local",
            "Cuenta Destino",
            "Estado",
        ]
        header_fill = self._get_header_fill(project)
        header_font = self._get_header_font()
        border = self._get_thin_border()

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # Data rows
        row = start_row + 1
        total_importe = Decimal("0")
        total_gastos = Decimal("0")
        total_neto = Decimal("0")

        for transfer in project.transfers:
            ws.cell(row=row, column=1, value=transfer.numero_display).border = border

            # Dates
            for col, fecha in [
                (2, transfer.fecha_peticion),
                (3, transfer.fecha_emision),
                (4, transfer.fecha_recepcion),
            ]:
                cell = ws.cell(
                    row=row,
                    column=col,
                    value=fecha.strftime("%d/%m/%Y") if fecha else "-",
                )
                cell.border = border

            # Amounts
            importe = transfer.importe_euros or Decimal("0")
            gastos = transfer.gastos_transferencia or Decimal("0")
            neto = transfer.importe_neto

            for col, value in [(5, importe), (6, gastos), (7, neto)]:
                cell = ws.cell(row=row, column=col, value=float(value))
                cell.number_format = '#,##0.00 "€"'
                cell.border = border

            # Exchange rate and local currency
            tipo_cambio = transfer.tipo_cambio_local or transfer.tipo_cambio_intermedio
            if tipo_cambio:
                cell = ws.cell(row=row, column=8, value=float(tipo_cambio))
                cell.number_format = "0.000000"
            else:
                ws.cell(row=row, column=8, value="-")
            ws[f"H{row}"].border = border

            ws.cell(row=row, column=9, value=transfer.moneda_local or "-").border = border

            if transfer.importe_moneda_local:
                cell = ws.cell(row=row, column=10, value=float(transfer.importe_moneda_local))
                cell.number_format = "#,##0.00"
            else:
                ws.cell(row=row, column=10, value="-")
            ws[f"J{row}"].border = border

            ws.cell(row=row, column=11, value=transfer.cuenta_destino or "-").border = border
            ws.cell(row=row, column=12, value=transfer.estado.value.replace("_", " ").title()).border = border

            total_importe += importe
            total_gastos += gastos
            total_neto += neto
            row += 1

        # Total row
        row += 1
        ws.cell(row=row, column=4, value="TOTALES:").font = Font(bold=True)

        for col, value in [(5, total_importe), (6, total_gastos), (7, total_neto)]:
            cell = ws.cell(row=row, column=col, value=float(value))
            cell.font = Font(bold=True)
            cell.number_format = '#,##0.00 "€"'

        self._auto_column_widths(ws)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"relacion_transferencias_{project.codigo_contable}_{timestamp}.xlsx"

        return buffer, filename

    def _generate_ficha_proyecto(
        self, project: Project, periodo: str | None = None
    ) -> tuple[BytesIO, str]:
        """Generate Ficha del Proyecto - Executive summary."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Ficha Proyecto"

        header_fill = self._get_header_fill(project)
        header_font = self._get_header_font()

        # Title
        ws["A1"] = "FICHA DEL PROYECTO"
        ws["A1"].font = self._get_title_font()
        ws.merge_cells("A1:D1")

        # Section: Identification
        row = 3
        ws.cell(row=row, column=1, value="IDENTIFICACION").font = Font(bold=True)
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=1).font = header_font
        ws.merge_cells(f"A{row}:D{row}")

        info = [
            ("Codigo Contable", project.codigo_contable),
            ("Codigo Area", project.codigo_area),
            ("Titulo", project.titulo),
            ("Pais", project.pais),
            ("Sector", project.sector),
            ("Tipo", project.tipo.value.replace("_", " ").title()),
            ("Financiador", project.financiador.value),
            ("Estado", project.estado.value.replace("_", " ").title()),
        ]

        for label, value in info:
            row += 1
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)

        # Section: Dates
        row += 2
        ws.cell(row=row, column=1, value="CRONOGRAMA").font = Font(bold=True)
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=1).font = header_font
        ws.merge_cells(f"A{row}:D{row}")

        dates = [
            ("Fecha Inicio", project.fecha_inicio.strftime("%d/%m/%Y")),
            ("Fecha Finalizacion", project.fecha_finalizacion.strftime("%d/%m/%Y")),
            (
                "Fecha Justificacion",
                project.fecha_justificacion.strftime("%d/%m/%Y")
                if project.fecha_justificacion
                else "No definida",
            ),
            ("Ampliado", "Si" if project.ampliado else "No"),
        ]

        for label, value in dates:
            row += 1
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)

        # Section: Financial
        row += 2
        ws.cell(row=row, column=1, value="FINANCIACION").font = Font(bold=True)
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=1).font = header_font
        ws.merge_cells(f"A{row}:D{row}")

        # Calculate totals
        total_aprobado = sum(bl.aprobado or Decimal("0") for bl in project.budget_lines)
        total_ejecutado = sum(
            e.cantidad_imputable
            for e in project.expenses
            if e.estado in (EstadoGasto.validado, EstadoGasto.justificado)
        )
        total_transferido = sum(
            t.importe_euros for t in project.transfers if t.estado.value in ("recibida", "cerrada")
        )

        row += 1
        ws.cell(row=row, column=1, value="Subvencion").font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=float(project.subvencion))
        cell.number_format = '#,##0.00 "€"'

        row += 1
        ws.cell(row=row, column=1, value="Presupuesto Aprobado").font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=float(total_aprobado))
        cell.number_format = '#,##0.00 "€"'

        row += 1
        ws.cell(row=row, column=1, value="Total Ejecutado").font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=float(total_ejecutado))
        cell.number_format = '#,##0.00 "€"'

        row += 1
        ws.cell(row=row, column=1, value="Total Transferido").font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=float(total_transferido))
        cell.number_format = '#,##0.00 "€"'

        if total_aprobado > 0:
            row += 1
            ws.cell(row=row, column=1, value="% Ejecucion").font = Font(bold=True)
            pct = float(total_ejecutado / total_aprobado)
            cell = ws.cell(row=row, column=2, value=pct)
            cell.number_format = "0.00%"

        # Section: ODS
        if project.ods_objetivos:
            row += 2
            ws.cell(row=row, column=1, value="ODS").font = Font(bold=True)
            ws.cell(row=row, column=1).fill = header_fill
            ws.cell(row=row, column=1).font = header_font
            ws.merge_cells(f"A{row}:D{row}")

            for ods in project.ods_objetivos:
                row += 1
                ws.cell(row=row, column=1, value=f"ODS {ods.numero}")
                ws.cell(row=row, column=2, value=ods.nombre)

        # Section: Metadata
        row += 2
        ws.cell(row=row, column=1, value="METADATOS").font = Font(bold=True)
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=1).font = header_font
        ws.merge_cells(f"A{row}:D{row}")

        row += 1
        ws.cell(row=row, column=1, value="Generado").font = Font(bold=True)
        ws.cell(row=row, column=2, value=datetime.now().strftime("%d/%m/%Y %H:%M"))

        self._auto_column_widths(ws)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"ficha_proyecto_{project.codigo_contable}_{timestamp}.xlsx"

        return buffer, filename

    def _generate_informe_tecnico(
        self, project: Project, periodo: str | None = None
    ) -> tuple[BytesIO, str]:
        """Generate Informe Tecnico Mensual."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Informe Tecnico"

        header_fill = self._get_header_fill(project)
        header_font = self._get_header_font()

        start_row = self._add_project_header(ws, project, "INFORME TECNICO MENSUAL")
        if periodo:
            ws.cell(row=start_row - 1, column=1, value=f"Periodo: {periodo}").font = Font(bold=True)

        row = start_row + 1

        # Activities section
        ws.cell(row=row, column=1, value="ACTIVIDADES REALIZADAS").fill = header_fill
        ws.cell(row=row, column=1).font = header_font
        ws.merge_cells(f"A{row}:G{row}")

        row += 1
        headers = ["Resultado", "Actividad", "Estado", "% Avance", "Observaciones"]
        border = self._get_thin_border()

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.border = border

        row += 1
        if project.logical_framework:
            for so in project.logical_framework.specific_objectives:
                for result in so.results:
                    for activity in result.activities:
                        ws.cell(row=row, column=1, value=result.codigo).border = border
                        ws.cell(row=row, column=2, value=activity.descripcion).border = border
                        ws.cell(row=row, column=3, value=activity.estado.value.replace("_", " ").title()).border = border

                        pct_cell = ws.cell(row=row, column=4, value=float(activity.porcentaje_avance or 0) / 100)
                        pct_cell.number_format = "0%"
                        pct_cell.border = border

                        ws.cell(row=row, column=5, value=activity.observaciones or "").border = border
                        row += 1

        # Indicators section
        row += 2
        ws.cell(row=row, column=1, value="INDICADORES").fill = header_fill
        ws.cell(row=row, column=1).font = header_font
        ws.merge_cells(f"A{row}:G{row}")

        row += 1
        headers = ["Indicador", "Meta", "Logrado", "% Cumplimiento", "Fuente"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.border = border

        row += 1
        if project.logical_framework:
            for so in project.logical_framework.specific_objectives:
                for result in so.results:
                    for indicator in result.indicators:
                        ws.cell(row=row, column=1, value=indicator.descripcion).border = border
                        ws.cell(row=row, column=2, value=float(indicator.meta or 0)).border = border
                        ws.cell(row=row, column=3, value=float(indicator.logrado or 0)).border = border

                        if indicator.meta and indicator.meta > 0:
                            pct = float(indicator.logrado or 0) / float(indicator.meta)
                            pct_cell = ws.cell(row=row, column=4, value=pct)
                            pct_cell.number_format = "0%"
                        else:
                            ws.cell(row=row, column=4, value="-")
                        ws[f"D{row}"].border = border

                        ws.cell(row=row, column=5, value=indicator.fuente_verificacion or "").border = border
                        row += 1

        self._auto_column_widths(ws)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        periodo_suffix = f"_{periodo}" if periodo else ""
        filename = f"informe_tecnico_{project.codigo_contable}{periodo_suffix}_{timestamp}.xlsx"

        return buffer, filename

    def _generate_informe_economico(
        self, project: Project, periodo: str | None = None
    ) -> tuple[BytesIO, str]:
        """Generate Informe Economico."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Informe Economico"

        start_row = self._add_project_header(ws, project, "INFORME ECONOMICO")

        header_fill = self._get_header_fill(project)
        header_font = self._get_header_font()
        border = self._get_thin_border()

        row = start_row

        # Summary section
        ws.cell(row=row, column=1, value="RESUMEN FINANCIERO").fill = header_fill
        ws.cell(row=row, column=1).font = header_font
        ws.merge_cells(f"A{row}:D{row}")

        # Calculate totals
        total_aprobado = sum(bl.aprobado or Decimal("0") for bl in project.budget_lines)
        total_ejecutado_espana = sum(
            e.cantidad_imputable
            for e in project.expenses
            if e.ubicacion == UbicacionGasto.espana
            and e.estado in (EstadoGasto.validado, EstadoGasto.justificado)
        )
        total_ejecutado_terreno = sum(
            e.cantidad_imputable
            for e in project.expenses
            if e.ubicacion == UbicacionGasto.terreno
            and e.estado in (EstadoGasto.validado, EstadoGasto.justificado)
        )
        total_ejecutado = total_ejecutado_espana + total_ejecutado_terreno
        total_transferido = sum(t.importe_euros for t in project.transfers)
        total_recibido = sum(
            t.importe_euros for t in project.transfers if t.estado.value == "recibida"
        )

        summary = [
            ("Presupuesto Aprobado", total_aprobado),
            ("Ejecutado España", total_ejecutado_espana),
            ("Ejecutado Terreno", total_ejecutado_terreno),
            ("Total Ejecutado", total_ejecutado),
            ("Pendiente de Ejecutar", total_aprobado - total_ejecutado),
            ("Total Transferido", total_transferido),
            ("Total Recibido", total_recibido),
        ]

        for label, value in summary:
            row += 1
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            cell = ws.cell(row=row, column=2, value=float(value))
            cell.number_format = '#,##0.00 "€"'

        # Execution by budget line
        row += 2
        ws.cell(row=row, column=1, value="EJECUCION POR PARTIDA").fill = header_fill
        ws.cell(row=row, column=1).font = header_font
        ws.merge_cells(f"A{row}:D{row}")

        row += 1
        headers = ["Partida", "Aprobado", "Ejecutado", "Disponible", "% Ejec."]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.border = border

        row += 1
        for bl in project.budget_lines:
            aprobado = bl.aprobado or Decimal("0")
            ejecutado = sum(
                e.cantidad_imputable
                for e in bl.expenses
                if e.estado in (EstadoGasto.validado, EstadoGasto.justificado)
            )
            disponible = aprobado - ejecutado

            ws.cell(row=row, column=1, value=bl.name).border = border

            for col, value in [(2, aprobado), (3, ejecutado), (4, disponible)]:
                cell = ws.cell(row=row, column=col, value=float(value))
                cell.number_format = '#,##0.00 "€"'
                cell.border = border

            pct = float(ejecutado / aprobado) if aprobado > 0 else 0
            pct_cell = ws.cell(row=row, column=5, value=pct)
            pct_cell.number_format = "0%"
            pct_cell.border = border

            row += 1

        # Recent expenses
        row += 2
        ws.cell(row=row, column=1, value="ULTIMOS GASTOS REGISTRADOS").fill = header_fill
        ws.cell(row=row, column=1).font = header_font
        ws.merge_cells(f"A{row}:D{row}")

        row += 1
        headers = ["Fecha", "Concepto", "Importe", "Estado"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.border = border

        row += 1
        recent_expenses = sorted(project.expenses, key=lambda e: e.fecha_factura, reverse=True)[:10]
        for expense in recent_expenses:
            ws.cell(row=row, column=1, value=expense.fecha_factura.strftime("%d/%m/%Y")).border = border
            ws.cell(row=row, column=2, value=expense.concepto[:50]).border = border
            cell = ws.cell(row=row, column=3, value=float(expense.cantidad_imputable))
            cell.number_format = '#,##0.00 "€"'
            cell.border = border
            ws.cell(row=row, column=4, value=expense.estado.value.replace("_", " ").title()).border = border
            row += 1

        self._auto_column_widths(ws)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"informe_economico_{project.codigo_contable}_{timestamp}.xlsx"

        return buffer, filename
